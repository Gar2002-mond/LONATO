from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField, Q, Avg, OuterRef, Subquery, Value
from django.db.models.functions import TruncMonth, TruncDay, Coalesce, ExtractYear, ExtractMonth, ExtractDay
from django.utils import timezone
from datetime import timedelta, datetime, date
from django.http import HttpResponse, JsonResponse
from decimal import Decimal
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from ..models import (
    ChiffreAffaire, Depense, Collecteur, Agent, ActiviteCollecteur, 
    PenaliteCommission, Machine, TypeDepense
)
from ..forms import DateRangeForm

def _format_excel_sheet(ws, headers, header_row=1, data_start_row=2, column_widths=None, 
                        date_column_indices=None, currency_column_indices=None, 
                        number_column_indices=None, percentage_column_indices=None, 
                        bold_header=True, wrap_text_header=False,
                        title_text=None, filter_texts=None):
    """
    Applique le formatage à une feuille Excel, avec option pour titre et filtres.
    """
    current_row = 1
    
    # Ajout du titre si fourni
    if title_text:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        title_cell = ws.cell(row=current_row, column=1, value=title_text)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 1
    
    # Ajout des filtres si fournis
    if filter_texts:
        for f_text in filter_texts:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            filter_cell = ws.cell(row=current_row, column=1, value=f_text)
            filter_cell.font = Font(italic=True)
            filter_cell.alignment = Alignment(horizontal='center')
            current_row += 1
    
    # Ajustement des lignes d'en-tête et de données
    if current_row > 1:
        header_row = current_row
        data_start_row = header_row + 1

    # Écriture des en-têtes
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header_title)
        if bold_header:
            cell.font = Font(bold=True)
        if wrap_text_header:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Initialisation des listes si None
    if date_column_indices is None: 
        date_column_indices = []
    if currency_column_indices is None: 
        currency_column_indices = []
    if number_column_indices is None: 
        number_column_indices = []
    if percentage_column_indices is None: 
        percentage_column_indices = []

    # Ajustement des largeurs de colonnes
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        
        if column_widths and col_idx <= len(column_widths) and column_widths[col_idx-1] is not None:
            ws.column_dimensions[column_letter].width = column_widths[col_idx-1]
        else:
            # Calcul automatique de la largeur
            max_length = len(str(header))
            
            # Parcourir les données pour trouver la largeur maximale
            for row_idx_iter in range(data_start_row, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx_iter, column=col_idx).value
                if cell_value is not None:
                    if isinstance(cell_value, (datetime, date)):
                        cell_length = 12  # Format DD/MM/YYYY
                    elif isinstance(cell_value, (int, float, Decimal)):
                        if col_idx in currency_column_indices:
                            cell_length = len(f"{cell_value:,.0f} FCFA")
                        elif col_idx in percentage_column_indices:
                            cell_length = len(f"{cell_value:.2f}%")
                        else:
                            cell_length = len(f"{cell_value:,.0f}")
                    else:
                        cell_length = len(str(cell_value))
                    
                    if cell_length > max_length:
                        max_length = cell_length
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Application des formats
    for row_num in range(data_start_row, ws.max_row + 1):
        # Format des dates
        for col_idx in date_column_indices:
            cell = ws.cell(row=row_num, column=col_idx)
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = 'DD/MM/YYYY'
        
        # Format des devises
        for col_idx in currency_column_indices:
            cell = ws.cell(row=row_num, column=col_idx)
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = '#,##0 "FCFA"'
        
        # Format des nombres
        for col_idx in number_column_indices:
            if col_idx not in currency_column_indices:
                cell = ws.cell(row=row_num, column=col_idx)
                if isinstance(cell.value, (int, float, Decimal)):
                    cell.number_format = '#,##0'
        
        # Format des pourcentages
        for col_idx in percentage_column_indices:
            cell = ws.cell(row=row_num, column=col_idx)
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = '0.00"%"'

@login_required
def rapport_chiffre_affaire(request):
    """Rapport sur les chiffres d'affaires"""
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    machine_id = request.GET.get('machine')
    form = DateRangeForm(request.GET or None)
    
    # Valeurs par défaut si non spécifiées
    if not date_debut:
        date_debut = (timezone.now() - timedelta(days=30)).date().strftime('%Y-%m-%d')
    if not date_fin:
        date_fin = timezone.now().date().strftime('%Y-%m-%d')
    
    chiffres = ChiffreAffaire.objects.filter(
        date_chiffre_affaire__range=[date_debut, date_fin]
    )
    
    if machine_id:
        chiffres = chiffres.filter(machine_id=machine_id)
    
    # Statistiques globales
    stats = chiffres.aggregate(
        total_ventes=Coalesce(Sum('ventes_totales'), Decimal('0')),
        total_annulations=Coalesce(Sum('annulations_totales'), Decimal('0')),
        total_paiements=Coalesce(Sum('paiements_totaux'), Decimal('0')),
        total_solde=Coalesce(Sum('solde_total'), Decimal('0')),
        total_commission=Coalesce(Sum('montant_commission'), Decimal('0')),
    )
    
    # Chiffres par machine
    chiffres_par_machine = chiffres.values(
        'machine__numero_terminal', 
        'machine__nom_point_vente'
    ).annotate(
        total_ventes=Coalesce(Sum('ventes_totales'), Decimal('0')),
        total_solde=Coalesce(Sum('solde_total'), Decimal('0')),
        total_commission=Coalesce(Sum('montant_commission'), Decimal('0')),
        nombre_jours=Count('id')
    ).order_by('-total_solde')
    
    # Chiffres par jour
    chiffres_par_jour = chiffres.values('date_chiffre_affaire').annotate(
        total_ventes=Coalesce(Sum('ventes_totales'), Decimal('0')),
        total_solde=Coalesce(Sum('solde_total'), Decimal('0')),
        total_commission=Coalesce(Sum('montant_commission'), Decimal('0')),
        nombre_machines=Count('machine', distinct=True)
    ).order_by('date_chiffre_affaire')
    
    context = {
        'form': form,
        'stats': stats,
        'chiffres_par_machine': chiffres_par_machine,
        'chiffres_par_jour': chiffres_par_jour,
        'machines': Machine.objects.all(),
        'machine_id': machine_id,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    
    return render(request, 'gestloto/rapports/chiffre_affaire.html', context)

@login_required
def rapport_depense(request):
    """Rapport sur les dépenses avec filtres avancés"""
    # Récupération des paramètres
    periode = request.GET.get('periode', '30')
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    type_depense_id = request.GET.get('type')
    agent_id = request.GET.get('agent')
    groupby = request.GET.get('groupby', 'jour')
    
    # Gestion des dates selon la période
    today = timezone.now().date()
    
    if periode == 'custom' and date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
            periode_texte = f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        except ValueError:
            date_debut = today - timedelta(days=30)
            date_fin = today
            periode_texte = "30 derniers jours (dates invalides)"
    else:
        # Périodes prédéfinies
        jours = int(periode) if periode.isdigit() else 30
        date_debut = today - timedelta(days=jours)
        date_fin = today
        
        periode_map = {
            '7': '7 derniers jours',
            '30': '30 derniers jours',
            '90': '3 derniers mois',
            '180': '6 derniers mois',
            '365': 'Année en cours'
        }
        periode_texte = periode_map.get(periode, f"{jours} derniers jours")
    
    # Requête de base
    depenses = Depense.objects.filter(
        date_depense__range=[date_debut, date_fin]
    ).select_related('type_depense', 'agent')
    
    # Filtres supplémentaires
    if type_depense_id and type_depense_id.isdigit():
        depenses = depenses.filter(type_depense_id=int(type_depense_id))
    
    if agent_id and agent_id.isdigit():
        depenses = depenses.filter(agent_id=int(agent_id))
    
    # Calcul des statistiques globales
    stats = depenses.aggregate(
        total_depenses=Coalesce(Sum('montant'), Decimal('0')),
        nb_transactions=Count('id'),
    )
    
    # Moyenne par transaction
    moyenne_transaction = stats['total_depenses'] / stats['nb_transactions'] if stats['nb_transactions'] > 0 else Decimal('0')
    
    resume = {
        'total_depenses': stats['total_depenses'],
        'nb_transactions': stats['nb_transactions'],
        'moyenne_transaction': moyenne_transaction,
    }
    
    # Répartition par type
    repartition_par_type = depenses.values(
        'type_depense__libelle'
    ).annotate(
        montant_total=Coalesce(Sum('montant'), Decimal('0')),
        nb_transactions=Count('id')
    ).order_by('-montant_total')
    
    # Calculer les pourcentages
    total_pour_pourcentage = stats['total_depenses']
    for item in repartition_par_type:
        if total_pour_pourcentage > 0:
            item['pourcentage'] = (item['montant_total'] / total_pour_pourcentage) * 100
        else:
            item['pourcentage'] = 0
        item['libelle_display'] = dict(TypeDepense.LIBELLE_CHOICES).get(item['type_depense__libelle'], item['type_depense__libelle'])
    
    # Données détaillées selon le groupement
    details = []
    
    if groupby == 'jour':
        details = depenses.values('date_depense').annotate(
            montant_total=Coalesce(Sum('montant'), Decimal('0')),
            nb_transactions=Count('id')
        ).order_by('date_depense')
        for item in details:
            item['date'] = item['date_depense']
    
    elif groupby == 'semaine':
        from django.db.models.functions import ExtractWeek
        details = depenses.annotate(
            semaine=ExtractWeek('date_depense'),
            annee=ExtractYear('date_depense')
        ).values('semaine', 'annee').annotate(
            montant_total=Coalesce(Sum('montant'), Decimal('0')),
            nb_transactions=Count('id')
        ).order_by('annee', 'semaine')
    
    elif groupby == 'mois':
        details = depenses.annotate(
            mois=TruncMonth('date_depense')
        ).values('mois').annotate(
            montant_total=Coalesce(Sum('montant'), Decimal('0')),
            nb_transactions=Count('id')
        ).order_by('mois')
        
        # Ajouter les noms de mois
        for item in details:
            if item['mois']:
                item['mois_texte'] = item['mois'].strftime('%B')
                item['annee'] = item['mois'].year
    
    elif groupby == 'type':
        details = depenses.values(
            'type_depense__libelle'
        ).annotate(
            montant_total=Coalesce(Sum('montant'), Decimal('0')),
            nb_transactions=Count('id')
        ).order_by('-montant_total')
        
        for item in details:
            item['libelle'] = item['type_depense__libelle']
            item['libelle_display'] = dict(TypeDepense.LIBELLE_CHOICES).get(item['libelle'], item['libelle'])
    
    # Dernières dépenses pour le tableau détaillé
    dernieres_depenses = depenses.order_by('-date_depense', '-id')[:20]
    
    # Données pour les filtres
    types_depenses = TypeDepense.objects.all()
    agents = Agent.objects.all().order_by('nom_agent', 'prenom_agent')
    
    context = {
        'resume': resume,
        'repartition_par_type': repartition_par_type,
        'details': details,
        'dernieres_depenses': dernieres_depenses,
        'types_depenses': types_depenses,
        'agents': agents,
        'periode_texte': periode_texte,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'request': request,
    }
    
    return render(request, 'gestloto/rapports/depense.html', context)

@login_required
def rapport_collecteur(request):
    """Rapport sur les collecteurs avec filtres avancés"""
    # Récupération des paramètres
    periode = request.GET.get('periode', '30')
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    collecteur_id = request.GET.get('collecteur')
    tri = request.GET.get('tri', 'montant')
    
    # Gestion des dates selon la période
    today = timezone.now().date()
    
    if periode == 'custom' and date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
            periode_texte = f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        except ValueError:
            date_debut = today - timedelta(days=30)
            date_fin = today
            periode_texte = "30 derniers jours (dates invalides)"
    else:
        # Périodes prédéfinies
        jours = int(periode) if periode.isdigit() else 30
        date_debut = today - timedelta(days=jours)
        date_fin = today
        
        periode_map = {
            '7': '7 derniers jours',
            '30': '30 derniers jours',
            '90': '3 derniers mois',
            '180': '6 derniers mois',
            '365': 'Année en cours'
        }
        periode_texte = periode_map.get(periode, f"{jours} derniers jours")
    
    # Requête de base pour les activités
    activites_queryset = ActiviteCollecteur.objects.filter(
        date_activite__range=[date_debut, date_fin]
    ).select_related('collecteur')
    
    # Filtre par collecteur spécifique
    collecteur_selectionne = None
    if collecteur_id and collecteur_id.isdigit():
        try:
            collecteur_selectionne = Collecteur.objects.get(id=int(collecteur_id))
            activites_queryset = activites_queryset.filter(collecteur=collecteur_selectionne)
        except Collecteur.DoesNotExist:
            pass
    
    # Calcul des statistiques globales
    stats_globales = activites_queryset.aggregate(
        total_montant=Coalesce(Sum('montant_activite'), Decimal('0.00')),
        nb_activites=Count('id'),
        taux_commission_moyen=Coalesce(Avg('pourcentage_commission'), Decimal('0.00'))
    )
    
    # Calcul de la commission totale
    commission_expression = ExpressionWrapper(
        F('montant_activite') * F('pourcentage_commission') / 100,
        output_field=DecimalField(max_digits=15, decimal_places=2)
    )
    
    total_commission = activites_queryset.annotate(
        commission_calc=commission_expression
    ).aggregate(
        total=Coalesce(Sum('commission_calc'), Decimal('0.00'))
    )['total']
    
    # Compter les collecteurs actifs et total
    collecteurs_actifs = activites_queryset.values('collecteur').distinct().count()
    collecteurs_total = Collecteur.objects.count()
    
    resume = {
        'total_montant': stats_globales['total_montant'],
        'total_commission': total_commission,
        'nb_activites': stats_globales['nb_activites'],
        'nb_collecteurs_actifs': collecteurs_actifs,
        'nb_collecteurs_total': collecteurs_total,
        'taux_commission_moyen': stats_globales['taux_commission_moyen']
    }
    
    # Données par collecteur
    collecteurs_data = activites_queryset.values(
        'collecteur__id',
        'collecteur__nom_collecteur',
        'collecteur__prenom_collecteur', 
        'collecteur__reference_collecteur',
        'collecteur__pourcentage_commission'
    ).annotate(
        total_montant=Coalesce(Sum('montant_activite'), Decimal('0.00')),
        total_commission=Coalesce(Sum(commission_expression), Decimal('0.00')),
        nb_activites=Count('id')
    )
    
    # Tri des données
    if tri == 'commission':
        collecteurs_data = collecteurs_data.order_by('-total_commission', '-total_montant')
    elif tri == 'activites':
        collecteurs_data = collecteurs_data.order_by('-nb_activites', '-total_montant')
    else:  # tri == 'montant' (défaut)
        collecteurs_data = collecteurs_data.order_by('-total_montant', '-total_commission')
    
    # Détails du collecteur sélectionné
    collecteur_details = None
    collecteur_activites = None
    
    if collecteur_selectionne:
        activites_collecteur = activites_queryset.filter(collecteur=collecteur_selectionne)
        
        collecteur_details = activites_collecteur.aggregate(
            total_montant=Coalesce(Sum('montant_activite'), Decimal('0.00')),
            total_commission=Coalesce(Sum(commission_expression), Decimal('0.00')),
            nb_activites=Count('id')
        )
        
        collecteur_activites = activites_collecteur.order_by('-date_activite', '-id')[:50]
    
    # Liste de tous les collecteurs pour le filtre
    tous_collecteurs = Collecteur.objects.order_by('nom_collecteur', 'prenom_collecteur')
    
    context = {
        'resume': resume,
        'collecteurs_data': collecteurs_data,
        'collecteurs': tous_collecteurs,
        'collecteur_selectionne': collecteur_selectionne,
        'collecteur_details': collecteur_details,
        'collecteur_activites': collecteur_activites,
        'periode_texte': periode_texte,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'request': request,
    }
    
    # Si c'est une requête HTMX, ne retourner que le contenu
    if request.headers.get('HX-Request'):
        return render(request, 'gestloto/rapports/collecteur_content.html', context)
    
    return render(request, 'gestloto/rapports/collecteur.html', context)

@login_required
def rapport_agent(request):
    """Rapport sur les agents avec filtres avancés"""
    # Récupération des paramètres
    periode = request.GET.get('periode', '30')
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    agent_id = request.GET.get('agent')
    tri = request.GET.get('tri', 'ca')
    vue = request.GET.get('vue', 'ca')
    ca_min = request.GET.get('ca_min')
    
    # Gestion des dates selon la période
    today = timezone.now().date()
    
    if periode == 'custom' and date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
            periode_texte = f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        except ValueError:
            date_debut = today - timedelta(days=30)
            date_fin = today
            periode_texte = "30 derniers jours (dates invalides)"
    else:
        # Périodes prédéfinies
        jours = int(periode) if periode.isdigit() else 30
        date_debut = today - timedelta(days=jours)
        date_fin = today
        
        periode_map = {
            '7': '7 derniers jours',
            '30': '30 derniers jours',
            '90': '3 derniers mois',
            '180': '6 derniers mois',
            '365': 'Année en cours'
        }
        periode_texte = periode_map.get(periode, f"{jours} derniers jours")
    
    # Requêtes de base
    chiffres = ChiffreAffaire.objects.filter(
        date_chiffre_affaire__range=[date_debut, date_fin]
    )
    
    penalites = PenaliteCommission.objects.filter(
        date_penalite__range=[date_debut, date_fin],
        montant_penalite__gt=0
    )
    
    commissions = PenaliteCommission.objects.filter(
        date_commission__range=[date_debut, date_fin],
        montant_commission__gt=0
    )
    
    # Filtre par agent spécifique
    agent_selectionne = None
    if agent_id and agent_id.isdigit():
        try:
            agent_selectionne = Agent.objects.get(id=int(agent_id))
            chiffres = chiffres.filter(agent_enregistreur=agent_selectionne)
            penalites = penalites.filter(agent=agent_selectionne)
            commissions = commissions.filter(agent=agent_selectionne)
        except Agent.DoesNotExist:
            pass
    
    # Calcul des statistiques globales
    stats_globales_ca = chiffres.aggregate(
        total_ca=Coalesce(Sum('solde_total'), Decimal('0')),
        total_commission_ca=Coalesce(Sum('montant_commission'), Decimal('0')),
        nb_enregistrements=Count('id')
    )
    
    stats_globales_penalites = penalites.aggregate(
        total_penalites=Coalesce(Sum('montant_penalite'), Decimal('0')),
        nb_penalites=Count('id')
    )
    
    stats_globales_commissions = commissions.aggregate(
        total_commissions=Coalesce(Sum('montant_commission'), Decimal('0')),
        nb_commissions=Count('id')
    )
    
    # Compter les agents actifs et total
    agents_actifs_ca = chiffres.values('agent_enregistreur').distinct().count()
    agents_total = Agent.objects.count()
    machines_utilisees = chiffres.values('machine').distinct().count()
    
    resume = {
        'total_ca': stats_globales_ca['total_ca'],
        'total_commission': stats_globales_commissions['total_commissions'],
        'total_penalite': stats_globales_penalites['total_penalites'],
        'balance': stats_globales_commissions['total_commissions'] - stats_globales_penalites['total_penalites'],
        'nb_agents_actifs': agents_actifs_ca,
        'nb_agents_total': agents_total,
        'nb_machines': machines_utilisees
    }
    
    # Données par agent avec toutes les métriques
    agents_queryset = Agent.objects.all()
    
    # Filtre par CA minimum
    if ca_min:
        try:
            ca_min_value = Decimal(ca_min)
        except (ValueError, TypeError):
            ca_min_value = None
    else:
        ca_min_value = None
    
    # Annotation des agents avec leurs statistiques
    agents_data = []
    for agent in agents_queryset:
        # CA de l'agent
        agent_ca = chiffres.filter(agent_enregistreur=agent).aggregate(
            total_ca=Coalesce(Sum('solde_total'), Decimal('0')),
            nb_machines=Count('machine', distinct=True)
        )
        
        # Commissions de l'agent
        agent_commissions = commissions.filter(agent=agent).aggregate(
            total_commission=Coalesce(Sum('montant_commission'), Decimal('0'))
        )
        
        # Pénalités de l'agent
        agent_penalites = penalites.filter(agent=agent).aggregate(
            total_penalite=Coalesce(Sum('montant_penalite'), Decimal('0'))
        )
        
        total_ca = agent_ca['total_ca']
        total_commission = agent_commissions['total_commission']
        total_penalite = agent_penalites['total_penalite']
        balance = total_commission - total_penalite
        
        # Filtre par CA minimum
        if ca_min_value and total_ca < ca_min_value:
            continue
        
        agent_data = {
            'id': agent.id,
            'nom_agent': agent.nom_agent,
            'prenom_agent': agent.prenom_agent,
            'reference_agent': agent.reference_agent,
            'total_ca': total_ca,
            'total_commission': total_commission,
            'total_penalite': total_penalite,
            'balance': balance,
            'nb_machines': agent_ca['nb_machines']
        }
        
        agents_data.append(agent_data)
    
    # Tri des données
    if tri == 'ca':
        agents_data.sort(key=lambda x: x['total_ca'], reverse=True)
    elif tri == 'ca_asc':
        agents_data.sort(key=lambda x: x['total_ca'])
    elif tri == 'commission':
        agents_data.sort(key=lambda x: x['total_commission'], reverse=True)
    elif tri == 'commission_asc':
        agents_data.sort(key=lambda x: x['total_commission'])
    elif tri == 'penalite':
        agents_data.sort(key=lambda x: x['total_penalite'], reverse=True)
    elif tri == 'balance':
        agents_data.sort(key=lambda x: x['balance'], reverse=True)
    elif tri == 'balance_asc':
        agents_data.sort(key=lambda x: x['balance'])
    elif tri == 'machines':
        agents_data.sort(key=lambda x: x['nb_machines'], reverse=True)
    elif tri == 'nom':
        agents_data.sort(key=lambda x: (x['nom_agent'], x['prenom_agent']))
    
    # Détails de l'agent sélectionné
    agent_details = None
    agent_chiffres_affaire = None
    agent_commissions = None
    agent_penalites = None
    agent_machines = None
    
    if agent_selectionne:
        # Statistiques détaillées
        agent_details = {
            'total_ca': chiffres.filter(agent_enregistreur=agent_selectionne).aggregate(
                total=Coalesce(Sum('solde_total'), Decimal('0'))
            )['total'],
            'total_commission': commissions.filter(agent=agent_selectionne).aggregate(
                total=Coalesce(Sum('montant_commission'), Decimal('0'))
            )['total'],
            'total_penalite': penalites.filter(agent=agent_selectionne).aggregate(
                total=Coalesce(Sum('montant_penalite'), Decimal('0'))
            )['total']
        }
        agent_details['balance'] = agent_details['total_commission'] - agent_details['total_penalite']
        
        # Données détaillées selon la vue
        if vue == 'ca' or not vue:
            agent_chiffres_affaire = chiffres.filter(
                agent_enregistreur=agent_selectionne
            ).select_related('machine').order_by('-date_chiffre_affaire')[:20]
        elif vue == 'commissions':
            agent_commissions = commissions.filter(
                agent=agent_selectionne
            ).select_related('machine').order_by('-date_commission')[:20]
        elif vue == 'penalites':
            agent_penalites = penalites.filter(
                agent=agent_selectionne
            ).select_related('machine').order_by('-date_penalite')[:20]
        
        # Machines gérées par l'agent
        agent_machines = Machine.objects.filter(agent=agent_selectionne)
    
    # Données pour le graphique
    chart_data = {
        'labels': [],
        'ca': [],
        'commission': [],
        'balance': []
    }
    
    # Générer les données de graphique par jour sur la période
    current_date = date_debut
    while current_date <= date_fin:
        chart_data['labels'].append(current_date.strftime('%d/%m'))
        
        # CA du jour
        ca_jour = chiffres.filter(date_chiffre_affaire=current_date).aggregate(
            total=Coalesce(Sum('solde_total'), Decimal('0'))
        )['total']
        
        # Commissions du jour
        comm_jour = commissions.filter(date_commission=current_date).aggregate(
            total=Coalesce(Sum('montant_commission'), Decimal('0'))
        )['total']
        
        # Pénalités du jour
        pen_jour = penalites.filter(date_penalite=current_date).aggregate(
            total=Coalesce(Sum('montant_penalite'), Decimal('0'))
        )['total']
        
        chart_data['ca'].append(float(ca_jour))
        chart_data['commission'].append(float(comm_jour))
        chart_data['balance'].append(float(comm_jour - pen_jour))
        
        current_date += timedelta(days=1)
    
    # Liste de tous les agents pour le filtre
    tous_agents = Agent.objects.order_by('nom_agent', 'prenom_agent')
    
    context = {
        'resume': resume,
        'agents_data': agents_data,
        'agents': tous_agents,
        'agent_selectionne': agent_selectionne,
        'agent_details': agent_details,
        'agent_chiffres_affaire': agent_chiffres_affaire,
        'agent_commissions': agent_commissions,
        'agent_penalites': agent_penalites,
        'agent_machines': agent_machines,
        'chart_data': chart_data,
        'periode_texte': periode_texte,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'request': request,
    }
    
    # Standardisation sur HTMX uniquement
    if request.headers.get('HX-Request'):
        return render(request, 'gestloto/rapports/agent_content.html', context)
    
    return render(request, 'gestloto/rapports/agent.html', context)

@login_required
def export_depenses_report_view(request):
    """Export Excel des dépenses"""
    # Récupération des paramètres
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    type_depense_id_str = request.GET.get('type')
    agent_id_str = request.GET.get('agent')

    # Requête de base
    queryset = Depense.objects.select_related('type_depense', 'agent').order_by('-date_depense', '-id')

    # Filtrage par date
    if date_debut_str and date_fin_str:
        try:
            queryset = queryset.filter(date_depense__range=[date_debut_str, date_fin_str])
        except ValueError:
            pass
    elif date_debut_str:
        try:
            queryset = queryset.filter(date_depense__gte=date_debut_str)
        except ValueError:
            pass
    elif date_fin_str:
        try:
            queryset = queryset.filter(date_depense__lte=date_fin_str)
        except ValueError:
            pass

    # Filtrage par type de dépense
    if type_depense_id_str:
        try:
            type_depense_id = int(type_depense_id_str)
            if type_depense_id > 0:
                queryset = queryset.filter(type_depense_id=type_depense_id)
        except (ValueError, TypeError):
            pass

    # Filtrage par agent
    if agent_id_str:
        try:
            agent_id = int(agent_id_str)
            if agent_id > 0:
                queryset = queryset.filter(agent_id=agent_id)
        except (ValueError, TypeError):
            pass

    # Création du classeur Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename_date = timezone.localdate().strftime("%Y%m%d")
    response['Content-Disposition'] = f'attachment; filename="rapport_depenses_{filename_date}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dépenses"

    # En-têtes
    headers = ["Date", "Type", "Référence", "Description", "Agent concerné", "Montant (FCFA)"]
    
    # Remplissage des données
    data_rows = [headers]
    for depense in queryset:
        agent_details = "-"
        if depense.agent:
            agent_details = f"{depense.agent.nom_agent or ''} {depense.agent.prenom_agent or ''}".strip()
            if not agent_details:
                agent_details = "-"

        type_display = "-"
        if depense.type_depense:
            type_display = depense.type_depense.get_libelle_display()
        
        description = getattr(depense, 'description', '-') or "-"

        data_rows.append([
            depense.date_depense,
            type_display,
            depense.reference or "-",
            description,
            agent_details,
            depense.montant,
        ])

    # Ajout des données au worksheet
    for row in data_rows:
        ws.append(row)

    # Formatage
    _format_excel_sheet(
        ws, headers,
        date_column_indices=[1],
        currency_column_indices=[6],
        title_text="Rapport des Dépenses",
        filter_texts=[f"Période d'export: {timezone.now().strftime('%d/%m/%Y')}"]
    )

    wb.save(response)
    return response

@login_required
def export_collecteurs_report_view(request):
    """Export Excel des collecteurs"""
    # Récupération des paramètres
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    collecteur_id_str = request.GET.get('collecteur')
    tri = request.GET.get('tri')

    # Requête de base
    queryset = ActiviteCollecteur.objects.select_related('collecteur').all()

    # Filtrage par date
    if date_debut_str and date_fin_str:
        try:
            queryset = queryset.filter(date_activite__range=[date_debut_str, date_fin_str])
        except ValueError:
            pass
    elif date_debut_str:
        try:
            queryset = queryset.filter(date_activite__gte=date_debut_str)
        except ValueError:
            pass
    elif date_fin_str:
        try:
            queryset = queryset.filter(date_activite__lte=date_fin_str)
        except ValueError:
            pass

    # Filtrage par collecteur
    if collecteur_id_str:
        try:
            collecteur_id = int(collecteur_id_str)
            if collecteur_id > 0:
                queryset = queryset.filter(collecteur_id=collecteur_id)
        except (ValueError, TypeError):
            pass

    # Annotation pour la commission calculée
    commission_expression = ExpressionWrapper(
        F('montant_activite') * F('pourcentage_commission') / 100,
        output_field=DecimalField(max_digits=15, decimal_places=2)
    )
    queryset = queryset.annotate(calculated_commission=commission_expression)

    # Tri
    if tri == 'montant':
        queryset = queryset.order_by('-montant_activite', '-date_activite')
    elif tri == 'commission':
        queryset = queryset.order_by('-calculated_commission', '-date_activite')
    else:
        queryset = queryset.order_by('-date_activite', 'collecteur__nom_collecteur')

    # Création du classeur Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename_date = timezone.localdate().strftime("%Y%m%d")
    response['Content-Disposition'] = f'attachment; filename="rapport_collecteurs_{filename_date}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activités Collecteurs"

    # En-têtes
    headers = [
        "Date Activité", "Collecteur", "Référence Reçu", 
        "Montant Activité (FCFA)", "% Commission", 
        "Commission Calculée (FCFA)", "Observations"
    ]

    # Remplissage des données
    data_rows = [headers]
    for activite in queryset:
        collecteur_info = "-"
        if activite.collecteur:
            nom = activite.collecteur.nom_collecteur or ""
            prenom = activite.collecteur.prenom_collecteur or ""
            ref = activite.collecteur.reference_collecteur or ""
            collecteur_info = f"{nom} {prenom} ({ref})".strip()
            if collecteur_info == "()":
                collecteur_info = ref or "-"

        data_rows.append([
            activite.date_activite,
            collecteur_info,
            activite.reference_recu or "-",
            activite.montant_activite,
            activite.pourcentage_commission,
            activite.calculated_commission,
            activite.observations or "-",
        ])

    # Ajout des données au worksheet
    for row in data_rows:
        ws.append(row)

    # Formatage
    _format_excel_sheet(
        ws, headers,
        date_column_indices=[1],
        currency_column_indices=[4, 6],
        percentage_column_indices=[5],
        title_text="Rapport des Collecteurs",
        filter_texts=[f"Période d'export: {timezone.now().strftime('%d/%m/%Y')}"]
    )

    wb.save(response)
    return response

@login_required
def export_agents_report_view(request):
    """Export Excel des agents"""
    # Récupération des paramètres
    periode = request.GET.get('periode', '30')
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    agent_id_str = request.GET.get('agent')
    tri = request.GET.get('tri', 'ca')
    ca_min = request.GET.get('ca_min')
    
    # Gestion des dates
    today = timezone.now().date()
    
    if periode == 'custom' and date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            date_debut = today - timedelta(days=30)
            date_fin = today
    else:
        jours = int(periode) if periode.isdigit() else 30
        date_debut = today - timedelta(days=jours)
        date_fin = today
    
    # Requêtes de base
    chiffres = ChiffreAffaire.objects.filter(
        date_chiffre_affaire__range=[date_debut, date_fin]
    )
    
    penalites = PenaliteCommission.objects.filter(
        date_penalite__range=[date_debut, date_fin],
        montant_penalite__gt=0
    )
    
    commissions = PenaliteCommission.objects.filter(
        date_commission__range=[date_debut, date_fin],
        montant_commission__gt=0
    )
    
    # Filtre par agent spécifique
    if agent_id_str:
        try:
            agent_id = int(agent_id_str)
            if agent_id > 0:
                chiffres = chiffres.filter(agent_enregistreur_id=agent_id)
                penalites = penalites.filter(agent_id=agent_id)
                commissions = commissions.filter(agent_id=agent_id)
        except (ValueError, TypeError):
            pass
    
    # Calcul des données par agent
    agents_queryset = Agent.objects.all()
    
    # Filtre par CA minimum
    if ca_min:
        try:
            ca_min_value = Decimal(ca_min)
        except (ValueError, TypeError):
            ca_min_value = None
    else:
        ca_min_value = None
    
    agents_data = []
    for agent in agents_queryset:
        # CA de l'agent
        agent_ca = chiffres.filter(agent_enregistreur=agent).aggregate(
            total_ca=Coalesce(Sum('solde_total'), Decimal('0')),
            nb_machines=Count('machine', distinct=True)
        )
        
        # Commissions de l'agent
        agent_commissions = commissions.filter(agent=agent).aggregate(
            total_commission=Coalesce(Sum('montant_commission'), Decimal('0'))
        )
        
        # Pénalités de l'agent
        agent_penalites = penalites.filter(agent=agent).aggregate(
            total_penalite=Coalesce(Sum('montant_penalite'), Decimal('0'))
        )
        
        total_ca = agent_ca['total_ca']
        total_commission = agent_commissions['total_commission']
        total_penalite = agent_penalites['total_penalite']
        balance = total_commission - total_penalite
        
        # Filtre par CA minimum
        if ca_min_value and total_ca < ca_min_value:
            continue
        
        agents_data.append({
            'reference_agent': agent.reference_agent,
            'nom_agent': agent.nom_agent,
            'prenom_agent': agent.prenom_agent,
            'total_ca': total_ca,
            'total_commission': total_commission,
            'total_penalite': total_penalite,
            'balance': balance,
            'nb_machines': agent_ca['nb_machines']
        })
    
    # Tri des données
    if tri == 'ca':
        agents_data.sort(key=lambda x: x['total_ca'], reverse=True)
    elif tri == 'ca_asc':
        agents_data.sort(key=lambda x: x['total_ca'])
    elif tri == 'commission':
        agents_data.sort(key=lambda x: x['total_commission'], reverse=True)
    elif tri == 'commission_asc':
        agents_data.sort(key=lambda x: x['total_commission'])
    elif tri == 'penalite':
        agents_data.sort(key=lambda x: x['total_penalite'], reverse=True)
    elif tri == 'balance':
        agents_data.sort(key=lambda x: x['balance'], reverse=True)
    elif tri == 'balance_asc':
        agents_data.sort(key=lambda x: x['balance'])
    elif tri == 'machines':
        agents_data.sort(key=lambda x: x['nb_machines'], reverse=True)
    elif tri == 'nom':
        agents_data.sort(key=lambda x: (x['nom_agent'], x['prenom_agent']))
    
    # Création du classeur Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename_date = timezone.localdate().strftime("%Y%m%d")
    response['Content-Disposition'] = f'attachment; filename="rapport_agents_{filename_date}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résumé Agents"

    # En-têtes
    headers = ["Réf. Agent", "Nom", "Prénom", "Chiffre d'Affaires", "Commissions Reçues", "Pénalités Subies", "Balance Nette", "Nb Machines"]
    
    # Remplissage des données
    data_rows = [headers]
    for agent in agents_data:
        data_rows.append([
            agent['reference_agent'],
            agent['nom_agent'],
            agent['prenom_agent'],
            agent['total_ca'],
            agent['total_commission'],
            agent['total_penalite'],
            agent['balance'],
            agent['nb_machines']
        ])

    # Ajout des données au worksheet
    for row in data_rows:
        ws.append(row)

    # Formatage
    _format_excel_sheet(
        ws, headers,
        currency_column_indices=[4, 5, 6, 7],
        number_column_indices=[8],
        title_text="Rapport des Agents",
        filter_texts=[
            f"Période: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}",
            f"Exporté le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ]
    )

    wb.save(response)
    return response

@login_required
def export_ca_report_view(request):
    # 1. Récupération et analyse des filtres
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    agent_id_str = request.GET.get('agent')
    machine_id_str = request.GET.get('machine')
    groupby = request.GET.get('groupby', 'jour') # défaut 'jour'

    # Dates par défaut et parsing
    try:
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date() if date_debut_str else (timezone.now() - timedelta(days=30)).date()
    except ValueError:
        date_debut = (timezone.now() - timedelta(days=30)).date()

    try:
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date() if date_fin_str else timezone.now().date()
    except ValueError:
        date_fin = timezone.now().date()

    # 2. Base Queryset for ChiffreAffaire
    ca_queryset = ChiffreAffaire.objects.filter(date_chiffre_affaire__range=[date_debut, date_fin])
    
    # Initialize filter display texts for Excel headers
    filter_display_texts = [f"Période: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"]

    if agent_id_str and agent_id_str != "0" and agent_id_str.isdigit():
        ca_queryset = ca_queryset.filter(agent_enregistreur_id=int(agent_id_str))
        try:
            agent_obj = Agent.objects.get(id=int(agent_id_str))
            filter_display_texts.append(f"Agent : {agent_obj.nom_agent} {agent_obj.prenom_agent}")
        except Agent.DoesNotExist:
            filter_display_texts.append(f"Agent ID : {agent_id_str} (non trouvé)")
            
    if machine_id_str and machine_id_str != "0" and machine_id_str.isdigit():
        ca_queryset = ca_queryset.filter(machine_id=int(machine_id_str))
        try:
            machine_obj = Machine.objects.get(id=int(machine_id_str))
            filter_display_texts.append(f"Machine : {machine_obj.numero_terminal} ({machine_obj.nom_point_vente})")
        except Machine.DoesNotExist:
            filter_display_texts.append(f"Machine ID : {machine_id_str} (non trouvée)")


    # Création du classeur Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename_date = timezone.localdate().strftime("%Y%m%d")
    response['Content-Disposition'] = f'attachment; filename="rapport_chiffre_affaires_{filename_date}.xlsx"'
    wb = openpyxl.Workbook()

    # Champs d'agrégation communs
    aggregations = {
        'total_ventes': Coalesce(Sum('ventes_totales'), Decimal('0.00'), output_field=DecimalField(max_digits=15, decimal_places=2)),
        'total_annulations': Coalesce(Sum('annulations_totales'), Decimal('0.00'), output_field=DecimalField(max_digits=15, decimal_places=2)),
        'total_paiements': Coalesce(Sum('paiements_totaux'), Decimal('0.00'), output_field=DecimalField(max_digits=15, decimal_places=2)),
        'total_solde': Coalesce(Sum('solde_total'), Decimal('0.00'), output_field=DecimalField(max_digits=15, decimal_places=2)),
        'total_commission': Coalesce(Sum('montant_commission'), Decimal('0.00'), output_field=DecimalField(max_digits=15, decimal_places=2)),
        'nombre_enregistrements': Count('id')
    }

    # --- Feuille 1: Résumé CA ---
    ws_summary = wb.active
    ws_summary.title = "Résumé CA"
    summary_data = ca_queryset.aggregate(**aggregations)
    
    summary_headers = ["Indicateur", "Valeur"]
    _format_excel_sheet(ws_summary, summary_headers, title_text="Résumé du Chiffre d'Affaires", filter_texts=filter_display_texts, column_widths=[30, 20])
    header_row_summary = ws_summary._current_row # Ligne après les titres/filtres + 1 pour les données
    
    ws_summary.append(["Total Ventes Brutes", summary_data['total_ventes']])
    ws_summary.append(["Total Annulations", summary_data['total_annulations']])
    ws_summary.append(["Total Paiements", summary_data['total_paiements']])
    ws_summary.append(["Solde Net Total", summary_data['total_solde']])
    ws_summary.append(["Total Commissions", summary_data['total_commission']])
    ws_summary.append(["Nombre d'enregistrements", summary_data['nombre_enregistrements']])
    
    # Formatage spécifique pour la feuille de résumé (les valeurs sont en colonne 2)
    for row_num in range(header_row_summary, ws_summary.max_row + 1):
        cell = ws_summary.cell(row=row_num, column=2)
        if isinstance(cell.value, (Decimal, float, int)):
            if ws_summary.cell(row=row_num, column=1).value == "Nombre d'enregistrements":
                cell.number_format = '#,##0'
            else:
                cell.number_format = '#,##0 "FCFA"'


    # --- Feuille 2: Détail CA groupé ---
    ws_detail = wb.create_sheet(title=f"Détail CA par {groupby.capitalize()}")
    detail_headers = []
    detail_data = []
    date_cols_detail, currency_cols_detail, num_cols_detail = [], [], []

    group_by_title_text = f"Détail du Chiffre d'Affaires groupé par {groupby.capitalize()}"

    if groupby == 'jour':
        detail_headers = ["Date", "Ventes", "Annulations", "Paiements", "Solde Net", "Commission", "Nb Enreg."]
        date_cols_detail = [1]
        currency_cols_detail = [2, 3, 4, 5, 6]
        num_cols_detail = [7]
        detail_data = ca_queryset.annotate(group_field=TruncDay('date_chiffre_affaire')) \
            .values('group_field').annotate(**aggregations).order_by('group_field')
        for row in detail_data: ws_detail.append([row['group_field'], row['total_ventes'], row['total_annulations'], row['total_paiements'], row['total_solde'], row['total_commission'], row['nombre_enregistrements']])

    elif groupby == 'semaine': # Simplifié: Année-Semaine
        detail_headers = ["Année", "Semaine", "Ventes", "Annulations", "Paiements", "Solde Net", "Commission", "Nb Enreg."]
        currency_cols_detail = [3, 4, 5, 6, 7]
        num_cols_detail = [1, 2, 8]
        # Note: Django's ExtractWeek might behave differently across database backends or might not be available.
        # Consider using TruncWeek if available and suitable, or custom logic for week calculation.
        # For simplicity, assuming ExtractWeek works as intended here.
        from django.db.models.functions import ExtractWeek # Ensure ExtractWeek is imported if used
        detail_data = ca_queryset.annotate(year=ExtractYear('date_chiffre_affaire'), week=ExtractWeek('date_chiffre_affaire')) \
            .values('year', 'week').annotate(**aggregations).order_by('year', 'week')
        for row in detail_data: ws_detail.append([row['year'], row['week'], row['total_ventes'], row['total_annulations'], row['total_paiements'], row['total_solde'], row['total_commission'], row['nombre_enregistrements']])
            
    elif groupby == 'mois':
        detail_headers = ["Mois", "Ventes", "Annulations", "Paiements", "Solde Net", "Commission", "Nb Enreg."]
        date_cols_detail = [1] # TruncMonth retourne une date
        currency_cols_detail = [2, 3, 4, 5, 6]
        num_cols_detail = [7]
        detail_data = ca_queryset.annotate(group_field=TruncMonth('date_chiffre_affaire')) \
            .values('group_field').annotate(**aggregations).order_by('group_field')
        for row in detail_data: ws_detail.append([row['group_field'], row['total_ventes'], row['total_annulations'], row['total_paiements'], row['total_solde'], row['total_commission'], row['nombre_enregistrements']])

    elif groupby == 'agent':
        detail_headers = ["Agent (Réf)", "Nom", "Prénom", "Ventes", "Annulations", "Paiements", "Solde Net", "Commission", "Nb Enreg."]
        currency_cols_detail = [4, 5, 6, 7, 8]
        num_cols_detail = [9]
        detail_data = ca_queryset.values('agent_enregistreur__reference_agent', 'agent_enregistreur__nom_agent', 'agent_enregistreur__prenom_agent') \
            .annotate(**aggregations).order_by('agent_enregistreur__nom_agent', 'agent_enregistreur__prenom_agent')
        for row in detail_data: ws_detail.append([row['agent_enregistreur__reference_agent'], row['agent_enregistreur__nom_agent'], row['agent_enregistreur__prenom_agent'], row['total_ventes'], row['total_annulations'], row['total_paiements'], row['total_solde'], row['total_commission'], row['nombre_enregistrements']])
            
    elif groupby == 'machine':
        detail_headers = ["Machine (N°)", "Point de Vente", "Ventes", "Annulations", "Paiements", "Solde Net", "Commission", "Nb Enreg."]
        currency_cols_detail = [3, 4, 5, 6, 7]
        num_cols_detail = [8]
        detail_data = ca_queryset.values('machine__numero_terminal', 'machine__nom_point_vente') \
            .annotate(**aggregations).order_by('machine__numero_terminal')
        for row in detail_data: ws_detail.append([row['machine__numero_terminal'], row['machine__nom_point_vente'], row['total_ventes'], row['total_annulations'], row['total_paiements'], row['total_solde'], row['total_commission'], row['nombre_enregistrements']])

    _format_excel_sheet(ws_detail, detail_headers, title_text=group_by_title_text, filter_texts=filter_display_texts,
                        date_column_indices=date_cols_detail, currency_column_indices=currency_cols_detail, number_column_indices=num_cols_detail, wrap_text_header=True)


    # --- Feuille 3: Top 5 Machines (par Solde Net) ---
    ws_top_machines = wb.create_sheet(title="Top 5 Machines")
    top_machines_headers = ["Machine (N°)", "Point de Vente", "Solde Net Total"]
    currency_cols_top_machines = [3]
    
    top_machines_data = ChiffreAffaire.objects.filter(date_chiffre_affaire__range=[date_debut, date_fin]) \
        .values('machine__numero_terminal', 'machine__nom_point_vente') \
        .annotate(total_solde_agg=Coalesce(Sum('solde_total'), Decimal('0.00'))) \
        .order_by('-total_solde_agg')[:5]
    
    for row in top_machines_data: ws_top_machines.append([row['machine__numero_terminal'], row['machine__nom_point_vente'], row['total_solde_agg']])
    _format_excel_sheet(ws_top_machines, top_machines_headers, title_text="Top 5 Machines par Solde Net", filter_texts=filter_display_texts,
                        currency_column_indices=currency_cols_top_machines, wrap_text_header=True)

    # --- Feuille 4: Top 5 Agents (par Solde Net généré) ---
    ws_top_agents = wb.create_sheet(title="Top 5 Agents")
    top_agents_headers = ["Agent (Réf)", "Nom", "Prénom", "Solde Net Total"]
    currency_cols_top_agents = [4]

    top_agents_data = ChiffreAffaire.objects.filter(date_chiffre_affaire__range=[date_debut, date_fin]) \
        .values('agent_enregistreur__reference_agent', 'agent_enregistreur__nom_agent', 'agent_enregistreur__prenom_agent') \
        .annotate(total_solde_agg=Coalesce(Sum('solde_total'), Decimal('0.00'))) \
        .order_by('-total_solde_agg')[:5]

    for row in top_agents_data: ws_top_agents.append([row['agent_enregistreur__reference_agent'], row['agent_enregistreur__nom_agent'], row['agent_enregistreur__prenom_agent'], row['total_solde_agg']])
    _format_excel_sheet(ws_top_agents, top_agents_headers, title_text="Top 5 Agents par Solde Net Généré", filter_texts=filter_display_texts,
                        currency_column_indices=currency_cols_top_agents, wrap_text_header=True)

    wb.save(response)
    return response